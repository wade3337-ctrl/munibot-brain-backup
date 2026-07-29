#!/usr/bin/env python3
"""
Generate a municipal bid cost-proposal spreadsheet from priced line items.

Usage:
  /opt/data/.venv/bin/python generate_bid_spreadsheet.py \
    --priced /path/to/priced.json \
    --signals /path/to/signals.json \
    --output /path/to/output.xlsx \
    --city "Long Beach" \
    --rfp "PW25-648" \
    --renewal-caps 7,7,7

Input: JSON array of priced line items with keys:
  description, unit, wca_2021, wca_2026_est, gsts_recommended,
  savings_vs_wca, savings_pct, rationale, below_floor_flag (optional)

Signals JSON must contain:
  price_buddy: {band: {avg_price, est_tph, est_hours, blended_floor, ...}}
  guardrails: {tph_target, escalation_rate, ...}

Produces: 4-tab Excel workbook
  1. Cost Proposal (the form to transcribe)
  2. Pricing Analysis (side-by-side comparison)
  3. Cost Floor & Weighting (Price Buddy + revenue by DBH band)
  4. Methodology (how the numbers were derived)

After generating, run verify_no_blanks.py to confirm zero empty cells.
"""
import argparse, json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def find_item(priced, search_term):
    """Find first item whose description contains search_term (case-insensitive)."""
    for item in priced:
        if search_term.upper() in item['description'].upper():
            return item
    return None


def generate(priced, signals, output_path, city, rfp_id, renewal_caps):
    wb = Workbook()

    # Styles
    header_font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    section_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    section_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    bold = Font(name='Calibri', bold=True)
    our_price_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    warn_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ============================================================
    # SHEET 1: COST PROPOSAL
    # ============================================================
    ws = wb.active
    ws.title = "Cost Proposal"
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40

    row = 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "ATTACHMENT AA — COST PROPOSAL"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 25
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = f"City of {city} — {rfp_id} Tree Trimming & Related Services"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "Great Scott Tree Service — Recommended Pricing (Bid Engine Output)"
    ws[f'A{row}'].font = Font(italic=True, size=10, color='666666')
    row += 2

    # Renewal caps
    caps_str = " / ".join(f"{c}%" for c in renewal_caps)
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = f"Renewal Price Increase Caps: {caps_str}"
    ws[f'A{row}'].font = Font(bold=True, size=10, color='C00000')
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = f"(Fill: ____% → {caps_str} for each renewal period)"
    ws[f'A{row}'].font = Font(italic=True, size=9, color='666666')
    row += 2

    def write_section(title):
        nonlocal row
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        row += 1

    def write_item(desc, unit, price):
        nonlocal row
        ws[f'A{row}'] = desc
        ws[f'B{row}'] = unit
        ws[f'C{row}'] = price
        ws[f'C{row}'].number_format = '$#,##0'
        ws[f'C{row}'].fill = our_price_fill
        ws[f'C{row}'].font = bold
        for col in 'ABCD':
            ws[f'{col}{row}'].border = thin_border
        row += 1

    def write_subheader(text):
        nonlocal row
        ws[f'A{row}'] = text
        ws[f'A{row}'].font = bold
        row += 1

    def safe_price(search):
        """Find item and return price, or '' if not found."""
        item = find_item(priced, search)
        return item['gsts_recommended'] if item else ''

    # 1. Hardwood Grid Trimming
    write_section("1. Hardwood Tree Grid Trimming")
    write_subheader("Full Prune Hardwood Tree")
    for label, search in [('0"-6" DBH','FULL PRUNE 0-6'),('7"-12" DBH','FULL PRUNE 7-12'),
        ('13"-18" DBH','FULL PRUNE 13-18'),('19"-24" DBH','FULL PRUNE 19-24'),
        ('24"-30" DBH','FULL PRUNE 24-30'),('Over 31" DBH','FULL PRUNE 31+')]:
        write_item(label, "EA", safe_price(search))
    write_subheader("Crown Raise/Clearance Prune Hardwood Tree")
    # CRITICAL: search for 'RAISE/CLEARANCE PRUNE 0-6' NOT 'RAISE/CLEARANCE 0-6'
    # Missing the word 'PRUNE' causes blank cells. See pitfall #8 in SKILL.md.
    for label, search in [('0"-6" DBH','RAISE/CLEARANCE PRUNE 0-6'),('7"-12" DBH','RAISE/CLEARANCE PRUNE 7-12'),
        ('13"-18" DBH','RAISE/CLEARANCE PRUNE 13-18'),('19"-24" DBH','RAISE/CLEARANCE PRUNE 19-24'),
        ('24"-30" DBH','RAISE/CLEARANCE PRUNE 24-30'),('Over 31" DBH','RAISE/CLEARANCE PRUNE 31+')]:
        write_item(label, "EA", safe_price(search))
    row += 1

    # 2. Palm Trimming
    write_section("2. Palm Tree Trimming")
    write_item("Prune Date Palm (Phoenix spp.)", "EA", safe_price('DATE PALM PRUNE'))
    write_item("Prune Fan Palm (Washingtonia spp.)", "EA", safe_price('FAN PALM PRUNE'))
    write_item("Prune All Other Palms", "EA", safe_price('PALM PRUNE, OTHER'))
    write_item("Clean Trunk: Date Palm (per ft.)", "FT", safe_price('TRUNK CLEAN: DATE'))
    write_item("Clean Trunk: Fan Palm (per ft.)", "FT", safe_price('TRUNK CLEAN: FAN'))
    row += 1

    # 3. Single Tree
    write_section("3. Single Tree Trimming")
    write_item("Full Prune Any Diameter Hardwood Tree", "EA", safe_price('SERVICE REQUEST PRUNE'))
    row += 1

    # 4. Removal
    write_section("4. Tree Removal")
    for label, search in [('0"-6" DBH','REMOVAL 0-6'),('7"-12" DBH','REMOVAL 7-12'),
        ('13"-18" DBH','REMOVAL 13-18'),('19"-24" DBH','REMOVAL 19-24'),
        ('24"-30" DBH','REMOVAL 24-30'),('Over 31" DBH','REMOVAL 31+')]:
        write_item(label, "EA", safe_price(search))
    row += 1

    # 5. Stump Grinding
    # CRITICAL: search for 'STUMP GRINDING 0-6' NOT 'STUMP 0-6'
    write_section("5. Stump Grinding")
    for label, search in [('0"-6" DBH','STUMP GRINDING 0-6'),('7"-12" DBH','STUMP GRINDING 7-12'),
        ('13"-18" DBH','STUMP GRINDING 13-18'),('19"-24" DBH','STUMP GRINDING 19-24'),
        ('24"-30" DBH','STUMP GRINDING 25-30'),('Over 31" DBH','STUMP GRINDING 31+')]:
        write_item(label, "EA", safe_price(search))
    row += 1

    # 6. Hourly
    write_section("6. General Hourly Labor Rates")
    write_item("Rate for One (1) Ground Person", "HR", safe_price('GROUND PERSON'))
    write_item("Rate for One (1) Equipment Operator", "HR", safe_price('EQUIPMENT OPERATOR'))
    write_item("Rate for One (1) Tree Trimmer", "HR", safe_price('TREE TRIMMER'))
    row += 1

    # 7. Day Rate
    write_section("7. Day-Rate Service Crew")
    write_item("Boom Truck Crew (8-hr day: trimmer + 2 ground)", "DAY", safe_price('DAY RATE'))
    row += 1

    # 8. Planting
    write_section("8. Tree Planting")
    write_item("15 Gallon", "EA", safe_price('15 GALLON'))
    write_item('24-inch Box', "EA", safe_price('24" BOX'))
    write_item('36-inch Box', "EA", safe_price('36" BOX'))
    write_item('48-inch Box', "EA", safe_price('48" BOX'))
    write_item("Fan Palm (10-30 ft. BTH)", "EA", safe_price('FAN PALM (10-30'))
    write_item("Fan Palm (per ft. BTH)", "FT", safe_price('FAN PALM (PER FT'))
    row += 1

    # 9. Watering
    write_section("9. Tree Watering")
    write_item("Water truck and Operator", "DAY", safe_price('WATERING'))
    row += 1

    # 10. Emergency
    write_section("10. Emergency Services")
    write_item("7AM-7PM Normal Business Hours", "HR", safe_price('DURING NORMAL'))
    write_item("7PM-7AM After-hours/weekends/holidays", "HR", safe_price('AFTER HOURS'))
    row += 1

    # 11. Support
    write_section("11. Support Services")
    write_item("ISA-Certified Arborist", "HR", safe_price('ARBORIST'))
    write_item("Pest Control Advisor", "HR", safe_price('PEST CONTROL'))
    write_item("Qualified Applicator", "HR", safe_price('QUALIFIED APPLICATOR'))

    # ============================================================
    # SHEET 2: PRICING ANALYSIS
    # ============================================================
    ws2 = wb.create_sheet("Pricing Analysis")
    headers = ['Section', 'Line Item', 'Unit', 'WCA 2021', 'WCA 2026 Est',
               'GSTS Rec.', 'Savings $', 'Savings %', 'WCA 2021', 'Rationale']
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws2.row_dimensions[1].height = 30
    for col, width in zip('ABCDEFGHIJ', [15, 45, 8, 12, 12, 12, 10, 10, 12, 40]):
        ws2.column_dimensions[col].width = width

    section_map = {
        'FULL PRUNE': '1a. Grid Trim', 'RAISE/CLEARANCE': '1b. Raise/Clear',
        'PALM': '2. Palms', 'SERVICE REQUEST': '3. Single Tree',
        'REMOVAL': '4. Removal', 'STUMP': '5. Stump Grind',
        'RATE FOR': '6. Hourly', 'DAY RATE': '7. Day Rate',
        'PLANT': '8. Planting', 'WATERING': '9. Watering',
        'EMERGENCY': '10. Emergency', 'ARBORIST': '11a. Arborist',
        'PEST CONTROL': '11b. PCA', 'QUALIFIED': '11c. QA',
    }

    def get_section(desc):
        for key, section in section_map.items():
            if key in desc.upper():
                return section
        return 'Other'

    row2 = 2
    for item in priced:
        ws2.cell(row=row2, column=1, value=get_section(item['description']))
        label = item['description'].title()
        ws2.cell(row=row2, column=2, value=label)
        ws2.cell(row=row2, column=3, value=item['unit'])
        ws2.cell(row=row2, column=4, value=item['wca_2021']).number_format = '$#,##0'
        ws2.cell(row=row2, column=5, value=item['wca_2026_est']).number_format = '$#,##0'
        ws2.cell(row=row2, column=6, value=item['gsts_recommended']).number_format = '$#,##0'
        ws2.cell(row=row2, column=6).fill = our_price_fill
        ws2.cell(row=row2, column=6).font = bold
        # Flag floor-enforced items in yellow
        if item.get('below_floor_flag'):
            ws2.cell(row=row2, column=6).fill = warn_fill
            ws2.cell(row=row2, column=2, value=label + ' ⚠ FLOOR')
        ws2.cell(row=row2, column=7, value=item['savings_vs_wca']).number_format = '$#,##0'
        ws2.cell(row=row2, column=8, value=item['savings_pct']/100).number_format = '0.0%'
        ws2.cell(row=row2, column=9, value=item.get('wca_2021', 0)).number_format = '$#,##0'
        ws2.cell(row=row2, column=10, value=item['rationale'])
        for col in range(1, 11):
            ws2.cell(row=row2, column=col).border = thin_border
        row2 += 1

    ws2.auto_filter.ref = f'A1:J{row2-1}'
    ws2.freeze_panes = 'A2'

    # ============================================================
    # SHEET 3: COST FLOOR & WEIGHTING
    # ============================================================
    ws3 = wb.create_sheet("Cost Floor & Weighting")
    for col, w in zip('ABCDEFG', [12, 12, 10, 12, 12, 10, 18]):
        ws3.column_dimensions[col].width = w

    row3 = 1
    ws3.merge_cells(f'A{row3}:G{row3}')
    ws3[f'A{row3}'] = "Cost Floor Analysis (Price Buddy — actual WorkOrderLine data, 2023+)"
    ws3[f'A{row3}'].font = header_font
    ws3[f'A{row3}'].fill = header_fill
    ws3[f'A{row3}'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[row3].height = 25
    row3 += 2

    for col, h in enumerate(['DBH Band', 'Tree Count', '% of Total', 'Avg Price',
                             'Avg TPH', 'Hours/Tree', 'Cost Floor ($130/hr)'], 1):
        cell = ws3.cell(row=row3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border
    row3 += 1

    # Tree counts from inventory — override per city if different
    inventory = {'0-6': 5606, '7-12': 15667, '13-18': 21590,
                 '19-24': 19054, '24-30': 22603, '31+': 2709}
    total = sum(inventory.values())
    price_buddy = signals.get('price_buddy', {})

    for band in ['0-6', '7-12', '13-18', '19-24', '24-30', '31+']:
        pb = price_buddy.get(band, {})
        ws3.cell(row=row3, column=1, value=band).font = bold
        ws3.cell(row=row3, column=2, value=inventory.get(band, 0)).number_format = '#,##0'
        ws3.cell(row=row3, column=3, value=inventory.get(band, 0)/total if total else 0).number_format = '0.0%'
        ws3.cell(row=row3, column=4, value=pb.get('avg_price', 0)).number_format = '$#,##0'
        ws3.cell(row=row3, column=5, value=pb.get('est_tph', 0)).number_format = '$#,##0'
        ws3.cell(row=row3, column=6, value=pb.get('est_hours', 0)).number_format = '0.000'
        ws3.cell(row=row3, column=7, value=pb.get('blended_floor', 0)).number_format = '$#,##0'
        ws3.cell(row=row3, column=7).fill = warn_fill
        for col in range(1, 8):
            ws3.cell(row=row3, column=col).border = thin_border
        row3 += 1

    ws3.cell(row=row3, column=1, value="TOTAL").font = bold
    ws3.cell(row=row3, column=2, value=total).number_format = '#,##0'
    ws3.cell(row=row3, column=2).font = bold
    for col in range(1, 8):
        ws3.cell(row=row3, column=col).border = thin_border
    row3 += 3

    # Weighted revenue
    ws3.merge_cells(f'A{row3}:G{row3}')
    ws3[f'A{row3}'] = "Weighted Annual Revenue Estimate (Grid Trimming Only)"
    ws3[f'A{row3}'].font = section_font
    ws3[f'A{row3}'].fill = section_fill
    row3 += 1

    for col, h in enumerate(['DBH Band', 'Tree Count', 'Our Price',
                             'Annual Revenue', 'WCA Revenue', 'We Save City'], 1):
        cell = ws3.cell(row=row3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border
    row3 += 1

    full_prune_keys = {'0-6': 'FULL PRUNE 0-6', '7-12': 'FULL PRUNE 7-12',
                       '13-18': 'FULL PRUNE 13-18', '19-24': 'FULL PRUNE 19-24',
                       '24-30': 'FULL PRUNE 24-30', '31+': 'FULL PRUNE 31+'}
    total_our_rev = 0
    total_wca_rev = 0
    for band, count in inventory.items():
        item = find_item(priced, full_prune_keys[band])
        if item:
            our_price = item['gsts_recommended']
            wca_price = item['wca_2026_est']
            our_rev = our_price * count
            wca_rev = wca_price * count
            total_our_rev += our_rev
            total_wca_rev += wca_rev
            ws3.cell(row=row3, column=1, value=band)
            ws3.cell(row=row3, column=2, value=count).number_format = '#,##0'
            ws3.cell(row=row3, column=3, value=our_price).number_format = '$#,##0'
            ws3.cell(row=row3, column=3).fill = our_price_fill
            ws3.cell(row=row3, column=4, value=our_rev).number_format = '$#,##0'
            ws3.cell(row=row3, column=5, value=wca_rev).number_format = '$#,##0'
            ws3.cell(row=row3, column=6, value=wca_rev - our_rev).number_format = '$#,##0'
        for col in range(1, 7):
            ws3.cell(row=row3, column=col).border = thin_border
        row3 += 1

    ws3.cell(row=row3, column=1, value="TOTAL").font = bold
    ws3.cell(row=row3, column=4, value=total_our_rev).number_format = '$#,##0'
    ws3.cell(row=row3, column=4).font = bold
    ws3.cell(row=row3, column=5, value=total_wca_rev).number_format = '$#,##0'
    ws3.cell(row=row3, column=5).font = bold
    ws3.cell(row=row3, column=6, value=total_wca_rev - total_our_rev).number_format = '$#,##0'
    ws3.cell(row=row3, column=6).font = bold

    # ============================================================
    # SHEET 4: METHODOLOGY
    # ============================================================
    ws4 = wb.create_sheet("Methodology")
    ws4.column_dimensions['A'].width = 100

    under_wca = sum(1 for p in priced if p['savings_vs_wca'] > 0)
    floor_items = sum(1 for p in priced if p.get('below_floor_flag'))
    avg_save = sum(p['savings_pct'] for p in priced if p['savings_vs_wca'] > 0)
    avg_save = avg_save / max(under_wca, 1)
    caps_str = " / ".join(f"{c}%" for c in renewal_caps)

    notes = [
        (f"{city} {rfp_id} — Bid Pricing Methodology", True, 14),
        ("", False, 11),
        ("SITUATION", True, 12),
        ("• Competitor (incumbent) bid escalated at 3.5%/yr to estimate their current pricing.", False, 11),
        (f"• {total:,} trees in the inventory across 6 DBH bands.", False, 11),
        ("• Contract term: 2 years base + 3 renewal periods.", False, 11),
        ("", False, 11),
        ("PRICING STRATEGY", True, 12),
        ("• Escalated incumbent's winning bid at 3.5%/year × 5 years = 18.8% total escalation.", False, 11),
        (f"• {under_wca}/{len(priced)} line items undercut the escalated estimate.", False, 11),
        ("• Discount tiers: 8% standard, 10% removals, 5% raise/clearance.", False, 11),
        ("• Cost floor enforcement: NEVER below $130/hr TPH on any labor rate.", False, 11),
        (f"• {floor_items} items floor-enforced (competitor below our cost — priced AT floor).", False, 11),
        (f"• Average savings on competitive items: {avg_save:.1f}%.", False, 11),
        ("", False, 11),
        ("COST FLOOR GUARDRAILS (all enforced)", True, 12),
        ("• Per-hour labor: $130/hr minimum (ground, operator, trimmer, arborist, PCA, QA)", False, 11),
        ("• Day rate crew: $3,120/day (3 persons × 8hr × $130)", False, 11),
        ("• Emergency day: $390/hr | Emergency night: $488/hr", False, 11),
        ("• Grid-trim per-tree: Price Buddy blended floor (128k+ WorkOrderLines)", False, 11),
        ("• Raise/clearance floor: 35% of prune floor", False, 11),
        ("• Removal floor: 2.5× prune floor | Stump floor: 35% of removal floor", False, 11),
        ("", False, 11),
        ("RENEWAL CAPS", True, 12),
        (f"• {caps_str} for the three renewal periods.", False, 11),
        ("", False, 11),
        ("DATA SOURCES", True, 12),
        ("1. Incumbent bid: Bid Results.pdf (prior contract cycle)", False, 11),
        ("2. Tree inventory: Tree Inventory_Pricing Worksheet.pdf", False, 11),
        ("3. GSTS current rates: TRIM IT LocationServiceTypes", False, 11),
        ("4. Cost floor: TRIM IT WorkOrderLines Price Buddy (128k+ records since 2023)", False, 11),
        ("5. Cost Proposal form: Appendix 1 / Attachment AA", False, 11),
        ("", False, 11),
        ("ITEMS NEEDING REVIEW", True, 12),
        ("• Floor-enforced items: competitor below our cost. Review with production team.", False, 11),
        ("• Day Rate: verify covers actual crew costs for 8-hour day.", False, 11),
        ("• Small-tree grid trim (0-6\", 7-12\"): competitor at/below our floor, high volume.", False, 11),
        ("", False, 11),
        ("NEXT STEPS", True, 12),
        ("1. Crew review: validate labor rates and day rate.", False, 11),
        ("2. Final review with Skipper before submission.", False, 11),
        ("3. Submit on the RFP's cost-proposal form (do NOT alter the form structure).", False, 11),
    ]

    for i, (text, is_bold, size) in enumerate(notes, 1):
        cell = ws4.cell(row=i, column=1, value=text)
        cell.font = Font(bold=is_bold, size=size)

    # Save
    wb.save(output_path)
    print(f"Spreadsheet saved: {output_path}")
    print(f"4 sheets: Cost Proposal | Pricing Analysis | Cost Floor & Weighting | Methodology")
    print(f"{len(priced)} items, {under_wca} under competitor, {floor_items} floor-enforced")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate bid cost-proposal spreadsheet')
    parser.add_argument('--priced', required=True, help='Path to priced JSON')
    parser.add_argument('--signals', required=True, help='Path to signals JSON')
    parser.add_argument('--output', required=True, help='Output .xlsx path')
    parser.add_argument('--city', default='Long Beach', help='City name')
    parser.add_argument('--rfp', default='PW25-648', help='RFP ID')
    parser.add_argument('--renewal-caps', default='7,7,7',
                        help='Comma-separated renewal cap percentages')
    args = parser.parse_args()

    with open(args.priced) as f:
        priced = json.load(f)
    with open(args.signals) as f:
        signals = json.load(f)

    renewal_caps = [int(x) for x in args.renewal_caps.split(',')]
    generate(priced, signals, args.output, args.city, args.rfp, renewal_caps)
