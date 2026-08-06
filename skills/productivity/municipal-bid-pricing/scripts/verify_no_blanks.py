#!/usr/bin/env python3
"""
Verify a generated cost-proposal spreadsheet has zero blank prices.

Generic: scans ALL sheets, finds every row with a populated item/description
column that also has a populated unit column, and flags any whose price/total
column is blank. Reports Crown Raise / Stump completeness only if those line
items exist in the sheet (some bids — e.g. Pomona — have Crown Raise but no
Stump Grinding; some have neither). Does not hard-fail on section counts.

Usage:
  /opt/data/.venv/bin/python verify_no_blanks.py /path/to/output.xlsx

Exits 0 if no blank prices found, exits 1 with the list if any found.
"""
import sys
from openpyxl import load_workbook

if len(sys.argv) < 2:
    print("Usage: verify_no_blanks.py <xlsx_path>")
    sys.exit(1)

path = sys.argv[1]
wb = load_workbook(path)

blanks = []
total_items = 0
crown_raise_priced = 0
stump_priced = 0

# Heuristic: the item/description is usually in column A (index 0), unit in
# column C or D, price in column D or E. Scan the first ~8 columns of every
# sheet and treat any row whose description column is a non-empty string AND
# whose unit column is a populated short string (ea/each/hr/day/ft/in/yd/...)
# as a "priced line item". Flag if the price column to the right of unit is blank.
UNIT_TOKENS = {"ea", "each", "hr", "hour", "hours", "day", "days", "ft", "feet",
               "in", "inch", "inches", "yd", "yard", "yards", "man hour", "per hour"}

def is_unit_token(v):
    if v is None: return False
    s = str(v).strip().lower()
    return s in UNIT_TOKENS

def is_description(v):
    if v is None or not isinstance(v, str): return False
    s = v.strip()
    if not s: return False
    # Skip headers, totals, section labels
    lower = s.lower()
    if lower.startswith(("total", "item", "qty", "description", "service level",
                         "level a", "level b", "extra work", "attachment",
                         "citywide", "great scott", "ifb")):
        return False
    if s.isupper() and len(s) < 30:  # ALL-CAPS section banner
        return False
    return True

for sn in wb.sheetnames:
    ws = wb[sn]
    # Skip analysis/meta sheets
    if any(k in sn.lower() for k in ["analysis", "comparison", "competitive",
                                      "reference", "notes", "top species",
                                      "volume", "cost floor"]):
        continue
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8):
        cells = [c.value for c in row]
        # find description cell (first string matching is_description)
        desc = None
        desc_idx = None
        for i, v in enumerate(cells):
            if is_description(v):
                desc = str(v).strip()
                desc_idx = i
                break
        if desc is None:
            continue
        # find a unit token to the right of the description
        unit_idx = None
        for i in range(desc_idx + 1, len(cells)):
            if is_unit_token(cells[i]):
                unit_idx = i
                break
        if unit_idx is None:
            continue
        # price is the first numeric cell to the right of unit
        price_idx = None
        for i in range(unit_idx + 1, len(cells)):
            v = cells[i]
            if isinstance(v, (int, float)):
                price_idx = i
                break
            # formula string also counts as "priced"
            if isinstance(v, str) and v.startswith("="):
                price_idx = i
                break
        total_items += 1
        lower_desc = desc.lower()
        if "crown raise" in lower_desc:
            crown_raise_priced += 1
        if "stump" in lower_desc and "removal" in lower_desc:
            stump_priced += 1
        if price_idx is None:
            blanks.append(f"[{sn}] row {row[0].row}: {desc[:60]}")

print(f"Sheets scanned: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
print(f"Total priced line items found: {total_items}")
if crown_raise_priced:
    print(f"Crown Raise items priced: {crown_raise_priced}")
if stump_priced:
    print(f"Stump Removal items priced: {stump_priced}")

if blanks:
    print(f"\n❌ {len(blanks)} BLANK PRICE(S):")
    for b in blanks:
        print(f"  - {b}")
    sys.exit(1)
else:
    print("\n✅ All line items priced. Zero blanks.")
    sys.exit(0)
